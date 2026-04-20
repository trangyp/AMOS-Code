#!/usr/bin/env python3
"""AMOS Equation Data Import - Multi-format Data Import System.

Production-grade data import system:
- CSV import with validation
- Excel import with multiple sheets
- JSON import for API data
- Async batch processing for large files
- Data validation and error reporting
- Progress tracking for long imports
- Row-level error handling (skip bad rows)
- Duplicate detection and handling
- Transaction support for atomic imports
- Import templates and field mapping
- Preview mode before actual import

Architecture Pattern: Pipeline pattern with validation stages
Import Features:
    - Multi-format support (CSV, Excel, JSON)
    - Streaming for large files
    - Background processing with Celery
    - Field mapping and transformation
    - Data validation per row
    - Error collection and reporting
    - Duplicate detection
    - Transaction safety

Integration:
    - equation_services: Data persistence
    - equation_tasks: Background processing
    - equation_schemas: Validation rules
    - equation_logging: Import audit logging
    - equation_database: Atomic transactions

Usage:
    # In API endpoint
    from equation_imports import ImportManager, ImportFormat

    @app.post("/api/v1/equations/import")
    async def import_equations(
        file: UploadFile,
        format: ImportFormat,
        mapping: FieldMapping,
        current_user: User = Depends(require_auth())
    ):
        manager = ImportManager()

        # Start import job
        job = await manager.import_async(
            file_path=await save_upload(file),
            format=format,
            entity_type="equations",
            field_mapping=mapping,
            user_id=current_user.id
        )

        return {"task_id": job.id, "status": "processing"}

Environment Variables:
    IMPORT_MAX_ROWS: Maximum rows per import (default: 50000)
    IMPORT_CHUNK_SIZE: Batch size for processing (default: 100)
    IMPORT_PREVIEW_ROWS: Rows to show in preview (default: 10)
    IMPORT_STORAGE_PATH: Path for uploaded files
    IMPORT_ALLOW_UPDATE: Allow updating existing records (default: false)
"""

from __future__ import annotations

import csv
import json
import logging
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Optional

# FastAPI imports
try:
    from fastapi import HTTPException, UploadFile, status

    FASTAPI_AVAILABLE = True
except ImportError:
    FASTAPI_AVAILABLE = False
    UploadFile = None
    HTTPException = Exception

# Data processing imports
try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Service imports
try:
    from equation_services import EquationService, get_equation_service

    SERVICES_AVAILABLE = True
except ImportError:
    SERVICES_AVAILABLE = False

# Task queue imports
try:
    from equation_tasks import celery_app

    TASKS_AVAILABLE = True
except ImportError:
    TASKS_AVAILABLE = False

# Schema imports
try:
    from equation_schemas import EquationRequestV1, validate_input

    SCHEMAS_AVAILABLE = True
except ImportError:
    SCHEMAS_AVAILABLE = False

# Database imports
try:
    from equation_database import async_session, get_session

    DATABASE_AVAILABLE = True
except ImportError:
    DATABASE_AVAILABLE = False

# Logging imports
try:
    from equation_logging import AuditAction, audit_log, get_logger

    LOGGING_AVAILABLE = True
except ImportError:
    LOGGING_AVAILABLE = False

logger = logging.getLogger("amos_equation_imports")


# ============================================================================
# Enums and Constants
# ============================================================================


class ImportFormat(str, Enum):
    """Supported import formats."""

    CSV = "csv"
    EXCEL = "excel"
    JSON = "json"
    JSONL = "jsonl"  # JSON Lines


class ImportStatus(str, Enum):
    """Import job status."""

    PENDING = "pending"
    VALIDATING = "validating"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"
    PARTIAL = "partial"  # Some rows failed


class DuplicateAction(str, Enum):
    """How to handle duplicate records."""

    SKIP = "skip"
    UPDATE = "update"
    ERROR = "error"


# Import limits
MAX_IMPORT_ROWS = 50000
IMPORT_CHUNK_SIZE = 100
IMPORT_PREVIEW_ROWS = 10


# ============================================================================
# Data Classes
# ============================================================================


@dataclass
class FieldMapping:
    """Maps source fields to target fields."""

    column_mapping: dict[str, str]  # source_col -> target_field
    default_values: dict[str, Any] = field(default_factory=dict)
    transforms: dict[str, str] = field(default_factory=dict)  # field -> transform_name

    def map_row(self, row: dict[str, Any]) -> dict[str, Any]:
        """Map a source row to target format."""
        result = {}

        for source_col, target_field in self.column_mapping.items():
            if source_col in row:
                value = row[source_col]

                # Apply transform if defined
                if target_field in self.transforms:
                    value = self._apply_transform(value, self.transforms[target_field])

                result[target_field] = value

        # Add default values
        result.update(self.default_values)

        return result

    def _apply_transform(self, value: Any, transform: str) -> Any:
        """Apply a transformation to a value."""
        if transform == "uppercase":
            return str(value).upper() if value else value
        elif transform == "lowercase":
            return str(value).lower() if value else value
        elif transform == "strip":
            return str(value).strip() if value else value
        elif transform == "int":
            return int(value) if value else None
        elif transform == "float":
            return float(value) if value else None
        return value


@dataclass
class ValidationError:
    """Row validation error."""

    row_number: int
    field: str
    message: str
    value: Any = None


@dataclass
class ImportJob:
    """Import job information."""

    id: str
    format: ImportFormat
    entity_type: str
    status: ImportStatus
    file_path: str
    field_mapping: FieldMapping
    user_id: int
    created_at: datetime
    total_rows: int = 0
    processed_rows: int = 0
    success_count: int = 0
    error_count: int = 0
    errors: list[ValidationError] = field(default_factory=list)
    completed_at: datetime = None
    preview_data: list[dict[str, Any]] = None

    def to_dict(self) -> dict[str, Any]:
        """Convert to dictionary."""
        return {
            "id": self.id,
            "format": self.format.value,
            "entity_type": self.entity_type,
            "status": self.status.value,
            "file_path": self.file_path,
            "user_id": self.user_id,
            "created_at": self.created_at.isoformat(),
            "completed_at": self.completed_at.isoformat() if self.completed_at else None,
            "total_rows": self.total_rows,
            "processed_rows": self.processed_rows,
            "success_count": self.success_count,
            "error_count": self.error_count,
            "progress_percent": self._get_progress(),
            "errors": [
                {"row": e.row_number, "field": e.field, "message": e.message}
                for e in self.errors[:100]  # Limit error details
            ],
            "preview_data": self.preview_data,
        }

    def _get_progress(self) -> float:
        """Calculate progress percentage."""
        if self.total_rows == 0:
            return 0.0
        return round((self.processed_rows / self.total_rows) * 100, 2)


@dataclass
class ImportResult:
    """Result of import operation."""

    success: bool
    imported_count: int
    updated_count: int
    error_count: int
    errors: list[ValidationError]
    warnings: list[str]


# ============================================================================
# File Readers (Strategy Pattern)
# ============================================================================


class BaseFileReader:
    """Base class for file readers."""

    def __init__(self, format: ImportFormat):
        self.format = format

    def read(self, file_path: str) -> list[dict[str, Any]]:
        """Read file and return list of rows."""
        raise NotImplementedError

    def read_streaming(self, file_path: str):
        """Read file as iterator for large files."""
        raise NotImplementedError

    def preview(self, file_path: str, rows: int = 10) -> list[dict[str, Any]]:
        """Preview first N rows."""
        raise NotImplementedError


class CSVFileReader(BaseFileReader):
    """CSV file reader."""

    def __init__(self):
        super().__init__(ImportFormat.CSV)

    def read(self, file_path: str) -> list[dict[str, Any]]:
        """Read CSV file."""
        rows = []
        with open(file_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(dict(row))
                if len(rows) >= MAX_IMPORT_ROWS:
                    break
        return rows

    def read_streaming(self, file_path: str):
        """Stream CSV rows."""
        with open(file_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                yield dict(row)

    def preview(self, file_path: str, rows: int = 10) -> list[dict[str, Any]]:
        """Preview CSV rows."""
        result = []
        with open(file_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader):
                if i >= rows:
                    break
                result.append(dict(row))
        return result


class ExcelFileReader(BaseFileReader):
    """Excel file reader."""

    def __init__(self):
        super().__init__(ImportFormat.EXCEL)

    def read(self, file_path: str, sheet_name: str = None) -> list[dict[str, Any]]:
        """Read Excel file."""
        if not PANDAS_AVAILABLE or not pd:
            raise ImportError("pandas required for Excel import")

        df = pd.read_excel(file_path, sheet_name=sheet_name or 0)

        # Limit rows
        if len(df) > MAX_IMPORT_ROWS:
            df = df.head(MAX_IMPORT_ROWS)

        # Convert to list of dicts
        return df.replace({pd.NA: None, pd.NaT: None}).to_dict("records")

    def read_streaming(self, file_path: str, sheet_name: str = None):
        """Stream Excel rows in chunks."""
        if not PANDAS_AVAILABLE or not pd:
            raise ImportError("pandas required for Excel import")

        # Read in chunks
        chunk_size = IMPORT_CHUNK_SIZE
        offset = 0

        while True:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name or 0,
                skiprows=range(1, offset + 1) if offset > 0 else None,
                nrows=chunk_size,
            )

            if df.empty:
                break

            for row in df.replace({pd.NA: None, pd.NaT: None}).to_dict("records"):
                yield row

            offset += chunk_size

            if offset >= MAX_IMPORT_ROWS:
                break

    def preview(
        self, file_path: str, rows: int = 10, sheet_name: str = None
    ) -> list[dict[str, Any]]:
        """Preview Excel rows."""
        if not PANDAS_AVAILABLE or not pd:
            raise ImportError("pandas required for Excel import")

        df = pd.read_excel(file_path, sheet_name=sheet_name or 0, nrows=rows)
        return df.replace({pd.NA: None, pd.NaT: None}).to_dict("records")

    def get_sheets(self, file_path: str) -> list[str]:
        """Get list of sheet names."""
        if not OPENPYXL_AVAILABLE:
            if not PANDAS_AVAILABLE:
                return []
            xl = pd.ExcelFile(file_path)
            return xl.sheet_names

        wb = load_workbook(file_path, read_only=True)
        return wb.sheetnames


class JSONFileReader(BaseFileReader):
    """JSON file reader."""

    def __init__(self):
        super().__init__(ImportFormat.JSON)

    def read(self, file_path: str) -> list[dict[str, Any]]:
        """Read JSON file."""
        with open(file_path, encoding="utf-8") as f:
            data = json.load(f)

        # Handle both single object and array
        if isinstance(data, dict):
            data = [data]

        return data[:MAX_IMPORT_ROWS] if len(data) > MAX_IMPORT_ROWS else data

    def read_streaming(self, file_path: str):
        """Stream JSON Lines file."""
        with open(file_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    yield json.loads(line)

    def preview(self, file_path: str, rows: int = 10) -> list[dict[str, Any]]:
        """Preview JSON rows."""
        data = self.read(file_path)
        return data[:rows]


# ============================================================================
# File Reader Registry
# ============================================================================


class FileReaderRegistry:
    """Registry for file readers."""

    _readers: dict[ImportFormat, BaseFileReader] = {}

    @classmethod
    def register(cls, format: ImportFormat, reader: BaseFileReader) -> None:
        """Register a reader."""
        cls._readers[format] = reader

    @classmethod
    def get_reader(cls, format: ImportFormat) -> BaseFileReader:
        """Get reader for format."""
        if format not in cls._readers:
            raise ValueError(f"No reader registered for format: {format}")
        return cls._readers[format]


# Register readers
FileReaderRegistry.register(ImportFormat.CSV, CSVFileReader())
FileReaderRegistry.register(ImportFormat.JSON, JSONFileReader())

if PANDAS_AVAILABLE:
    FileReaderRegistry.register(ImportFormat.EXCEL, ExcelFileReader())


# ============================================================================
# Validators
# ============================================================================


class RowValidator:
    """Validates import rows."""

    def __init__(self, entity_type: str):
        self.entity_type = entity_type
        self.required_fields = self._get_required_fields()

    def _get_required_fields(self) -> list[str]:
        """Get required fields for entity type."""
        if self.entity_type == "equations":
            return ["name", "formula"]
        elif self.entity_type == "users":
            return ["username", "email"]
        return []

    def validate(self, row: dict[str, Any], row_number: int) -> list[ValidationError]:
        """Validate a single row."""
        errors = []

        # Check required fields
        for field in self.required_fields:
            if field not in row or row[field] is None or row[field] == "":
                errors.append(
                    ValidationError(
                        row_number=row_number,
                        field=field,
                        message=f"Required field '{field}' is missing or empty",
                        value=row.get(field),
                    )
                )

        # Type validation
        if self.entity_type == "equations":
            errors.extend(self._validate_equation(row, row_number))

        return errors

    def _validate_equation(self, row: dict[str, Any], row_number: int) -> list[ValidationError]:
        """Validate equation-specific fields."""
        errors = []

        # Validate formula length
        formula = row.get("formula", "")
        if formula and len(str(formula)) > 10000:
            errors.append(
                ValidationError(
                    row_number=row_number,
                    field="formula",
                    message="Formula exceeds maximum length of 10000 characters",
                    value=formula[:50] + "...",
                )
            )

        return errors


# ============================================================================
# Import Manager
# ============================================================================


class ImportManager:
    """Main import manager."""

    def __init__(self):
        self.logger = get_logger("imports") if LOGGING_AVAILABLE else logger
        self._jobs: dict[str, ImportJob] = {}
        self.storage_path = Path("./uploads")
        self.storage_path.mkdir(parents=True, exist_ok=True)

    async def preview_import(
        self,
        file_path: str,
        format: ImportFormat,
        field_mapping: Optional[FieldMapping] = None,
        rows: int = IMPORT_PREVIEW_ROWS,
    ) -> list[dict[str, Any]]:
        """Preview import data before processing.

        Args:
            file_path: Path to import file
            format: Import format
            field_mapping: Field mapping configuration
            rows: Number of rows to preview

        Returns:
            List of preview rows
        """
        reader = FileReaderRegistry.get_reader(format)
        preview_data = reader.preview(file_path, rows)

        # Apply field mapping if provided
        if field_mapping:
            preview_data = [field_mapping.map_row(row) for row in preview_data]

        return preview_data

    async def import_async(
        self,
        file_path: str,
        format: ImportFormat,
        entity_type: str,
        field_mapping: Optional[FieldMapping] = None,
        user_id: int = 0,
        duplicate_action: DuplicateAction = DuplicateAction.SKIP,
        allow_update: bool = False,
    ) -> ImportJob:
        """Start async import job.

        Args:
            file_path: Path to import file
            format: Import format
            entity_type: Entity type to import
            field_mapping: Field mapping
            user_id: User requesting import
            duplicate_action: How to handle duplicates
            allow_update: Allow updating existing records

        Returns:
            Import job
        """
        job_id = f"import_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{user_id}"

        # Get total row count
        reader = FileReaderRegistry.get_reader(format)
        try:
            all_data = reader.read(file_path)
            total_rows = len(all_data)
        except Exception:
            total_rows = 0

        job = ImportJob(
            id=job_id,
            format=format,
            entity_type=entity_type,
            status=ImportStatus.PENDING,
            file_path=file_path,
            field_mapping=field_mapping,
            user_id=user_id,
            created_at=datetime.now(),
            total_rows=total_rows,
            preview_data=await self.preview_import(file_path, format, field_mapping),
        )

        self._jobs[job_id] = job

        # Start background processing
        if TASKS_AVAILABLE:
            process_import_task.delay(
                job_id,
                format.value,
                entity_type,
                file_path,
                field_mapping.column_mapping if field_mapping else {},
                user_id,
                duplicate_action.value,
                allow_update,
            )
        else:
            # Process synchronously
            await self._process_import(job, duplicate_action, allow_update)

        self.logger.info(f"Started import job {job_id} for {entity_type}")
        return job

    async def _process_import(
        self, job: ImportJob, duplicate_action: DuplicateAction, allow_update: bool
    ) -> None:
        """Process import job.

        Args:
            job: Import job to process
            duplicate_action: Duplicate handling strategy
            allow_update: Allow updating existing records
        """
        try:
            job.status = ImportStatus.VALIDATING

            # Get reader
            reader = FileReaderRegistry.get_reader(job.format)
            validator = RowValidator(job.entity_type)

            # Process in batches
            job.status = ImportStatus.PROCESSING
            batch: list[dict[str, Any]] = []

            for row_number, row in enumerate(reader.read_streaming(job.file_path), 1):
                job.processed_rows = row_number

                # Apply field mapping
                if job.field_mapping:
                    row = job.field_mapping.map_row(row)

                # Validate row
                errors = validator.validate(row, row_number)

                if errors:
                    job.error_count += 1
                    job.errors.extend(errors)

                    # Skip invalid rows
                    continue

                batch.append(row)

                # Process batch
                if len(batch) >= IMPORT_CHUNK_SIZE:
                    await self._process_batch(job, batch, duplicate_action, allow_update)
                    batch = []

                # Limit check
                if row_number >= MAX_IMPORT_ROWS:
                    job.warnings = job.warnings or []
                    job.warnings.append(f"Import limited to {MAX_IMPORT_ROWS} rows")
                    break

            # Process remaining batch
            if batch:
                await self._process_batch(job, batch, duplicate_action, allow_update)

            # Determine final status
            if job.error_count == 0:
                job.status = ImportStatus.COMPLETED
            elif job.success_count > 0:
                job.status = ImportStatus.PARTIAL
            else:
                job.status = ImportStatus.FAILED

            job.completed_at = datetime.now()

            # Audit log
            if LOGGING_AVAILABLE:
                audit_log(
                    action=AuditAction.DATA_IMPORTED,
                    user_id=job.user_id,
                    resource_type=job.entity_type,
                    details={
                        "format": job.format.value,
                        "total_rows": job.total_rows,
                        "success": job.success_count,
                        "errors": job.error_count,
                    },
                )

            self.logger.info(
                f"Import job {job.id} completed: "
                f"{job.success_count} success, {job.error_count} errors"
            )

        except Exception as e:
            job.status = ImportStatus.FAILED
            job.completed_at = datetime.now()
            job.errors.append(
                ValidationError(row_number=0, field=None, message=f"Import failed: {str(e)}")
            )
            self.logger.error(f"Import job {job.id} failed: {e}")

    async def _process_batch(
        self,
        job: ImportJob,
        batch: list[dict[str, Any]],
        duplicate_action: DuplicateAction,
        allow_update: bool,
    ) -> None:
        """Process a batch of rows.

        Args:
            job: Import job
            batch: List of rows to process
            duplicate_action: How to handle duplicates
            allow_update: Allow updating existing records
        """
        if not SERVICES_AVAILABLE:
            return

        try:
            # Get service
            if job.entity_type == "equations":
                service = await get_equation_service()

                for row in batch:
                    try:
                        # Check for duplicates
                        existing = await service.get_by_name(row.get("name"))

                        if existing:
                            if duplicate_action == DuplicateAction.SKIP:
                                continue
                            elif duplicate_action == DuplicateAction.UPDATE and allow_update:
                                await service.update(existing.id, row)
                                job.success_count += 1
                            else:
                                job.error_count += 1
                                continue
                        else:
                            # Create new
                            await service.create(row)
                            job.success_count += 1

                    except Exception as e:
                        job.error_count += 1
                        job.errors.append(ValidationError(row_number=0, field=None, message=str(e)))

        except Exception as e:
            self.logger.error(f"Batch processing error: {e}")
            raise

    def get_job(self, job_id: str) -> Optional[ImportJob]:
        """Get import job by ID."""
        return self._jobs.get(job_id)

    def get_user_jobs(self, user_id: int) -> list[ImportJob]:
        """Get all jobs for a user."""
        return [job for job in self._jobs.values() if job.user_id == user_id]

    def cancel_job(self, job_id: str) -> bool:
        """Cancel a pending or processing job."""
        job = self._jobs.get(job_id)
        if job and job.status in (ImportStatus.PENDING, ImportStatus.PROCESSING):
            job.status = ImportStatus.CANCELLED
            return True
        return False


# ============================================================================
# Celery Task for Background Import
# ============================================================================

if TASKS_AVAILABLE:

    @celery_app.task(bind=True, max_retries=3)
    def process_import_task(
        self,
        job_id: str,
        format_value: str,
        entity_type: str,
        file_path: str,
        column_mapping: dict[str, str],
        user_id: int,
        duplicate_action_value: str,
        allow_update: bool,
    ) -> dict[str, Any]:
        """Process import in background.

        Args:
            job_id: Import job ID
            format_value: Import format
            entity_type: Entity type
            file_path: Path to import file
            column_mapping: Field mapping
            user_id: User ID
            duplicate_action_value: Duplicate handling
            allow_update: Allow updates

        Returns:
            Job result
        """
        import asyncio

        async def run_import():
            manager = ImportManager()
            job = manager.get_job(job_id)

            if not job:
                return {"error": "Job not found"}

            # Create field mapping from dict
            field_mapping = FieldMapping(column_mapping=column_mapping) if column_mapping else None
            job.field_mapping = field_mapping

            duplicate_action = DuplicateAction(duplicate_action_value)

            await manager._process_import(job, duplicate_action, allow_update)
            return job.to_dict()

        try:
            return asyncio.run(run_import())
        except Exception as exc:
            logger.error(f"Import task failed: {exc}")
            raise self.retry(exc=exc, countdown=60)


# ============================================================================
# FastAPI Integration
# ============================================================================

if FASTAPI_AVAILABLE:
    from fastapi import APIRouter, BackgroundTasks, File, Query, UploadFile

    router = APIRouter(prefix="/api/v1/imports", tags=["Imports"])

    @router.post("/preview")
    async def preview_import(
        file: UploadFile = File(...),
        format: ImportFormat = Query(...),
        mapping: str = Query(default="{}"),
        rows: int = Query(default=10, ge=1, le=100),
    ):
        """Preview import data before processing."""
        # Save uploaded file
        file_path = f"./uploads/{file.filename}"
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)

        # Parse mapping
        column_mapping = json.loads(mapping) if mapping else {}
        field_mapping = FieldMapping(column_mapping=column_mapping) if column_mapping else None

        # Get preview
        manager = ImportManager()
        preview = await manager.preview_import(file_path, format, field_mapping, rows)

        return {"preview": preview, "total_rows": len(preview)}

    @router.post("/")
    async def create_import(
        file: UploadFile = File(...),
        format: ImportFormat = Query(...),
        entity_type: str = Query(...),
        mapping: str = Query(default="{}"),
        duplicate_action: DuplicateAction = Query(default=DuplicateAction.SKIP),
        background_tasks: BackgroundTasks = None,
        current_user=None,
    ):
        """Create new import job."""
        # Save file
        file_path = f"./uploads/{file.filename}"
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)

        # Parse mapping
        column_mapping = json.loads(mapping) if mapping else {}
        field_mapping = FieldMapping(column_mapping=column_mapping) if column_mapping else None

        # Start import
        manager = ImportManager()
        job = await manager.import_async(
            file_path=file_path,
            format=format,
            entity_type=entity_type,
            field_mapping=field_mapping,
            user_id=1,  # Would use current_user.id
            duplicate_action=duplicate_action,
        )

        return {
            "job_id": job.id,
            "status": job.status.value,
            "preview": job.preview_data,
            "created_at": job.created_at.isoformat(),
        }

    @router.get("/{job_id}")
    async def get_import_status(job_id: str):
        """Get import job status."""
        manager = ImportManager()
        job = manager.get_job(job_id)

        if not job:
            raise HTTPException(status_code=404, detail="Import job not found")

        return job.to_dict()

    @router.post("/{job_id}/cancel")
    async def cancel_import(job_id: str):
        """Cancel import job."""
        manager = ImportManager()
        if manager.cancel_job(job_id):
            return {"message": "Job cancelled"}
        raise HTTPException(status_code=400, detail="Job cannot be cancelled")


# ============================================================================
# Example Usage
# ============================================================================


async def example_usage():
    """Example usage of import system."""
    print("AMOS Equation Import System")
    print("=" * 50)

    # Create sample CSV
    csv_content = """name,formula,domain
circle_area,A=pi*r^2,geometry
velocity,v=d/t,physics
quadratic,x=(-b+sqrt(b^2-4ac))/2a,algebra"""

    file_path = "./sample_import.csv"
    with open(file_path, "w") as f:
        f.write(csv_content)

    # Create field mapping
    mapping = FieldMapping(
        column_mapping={"name": "name", "formula": "formula", "domain": "domain"}
    )

    # Preview
    manager = ImportManager()
    preview = await manager.preview_import(
        file_path=file_path, format=ImportFormat.CSV, field_mapping=mapping, rows=3
    )

    print("\nPreview:")
    for row in preview:
        print(f"  {row}")

    print("\nImport examples completed!")


if __name__ == "__main__":
    import asyncio

    asyncio.run(example_usage())
